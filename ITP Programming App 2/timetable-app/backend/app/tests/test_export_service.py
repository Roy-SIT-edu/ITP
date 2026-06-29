"""Tests for generated timetable exports."""

import csv
from io import StringIO

from openpyxl import load_workbook

from app.models.room import Room
from app.models.schedule_run import ScheduleRun
from app.models.scheduled_session import ScheduledSession
from app.models.session import Session
from app.models.time_slot import TimeSlot
from app.services.export_service import SYSTEM_TEMPLATE_COLUMNS, ExportService


def _add_scheduled_session(db_session) -> ScheduleRun:
    session = db_session.query(Session).filter_by(requirement_id="REQ-DEMO-001").one()
    session.custom_weeks = "1,2,3"
    session.combined_with_programmes = "ESE1109"
    room = db_session.query(Room).filter_by(room_code="LECT-01").one()
    slot = (
        db_session.query(TimeSlot)
        .filter_by(day="Wednesday", start_time="09:00", end_time="11:00", week_pattern="Weekly")
        .one()
    )
    run = ScheduleRun(status="COMPLETED")
    db_session.add(run)
    db_session.flush()
    db_session.add(
        ScheduledSession(
            schedule_run_id=run.id,
            session_id=session.id,
            room_id=room.id,
            time_slot_id=slot.id,
            staff_id=session.staff_id,
            day=slot.day,
            start_time=slot.start_time,
            end_time=slot.end_time,
            week_pattern=slot.week_pattern,
        )
    )
    db_session.commit()
    return run


def test_xlsx_export_uses_system_template_columns_and_values(db_session):
    run = _add_scheduled_session(db_session)

    workbook = load_workbook(ExportService().xlsx_buffer(db_session, run.id))
    sheet = workbook["Sheet1"]

    assert [sheet.cell(1, column).value for column in range(1, len(SYSTEM_TEMPLATE_COLUMNS) + 1)] == SYSTEM_TEMPLATE_COLUMNS
    assert [sheet.cell(2, column).value for column in range(1, len(SYSTEM_TEMPLATE_COLUMNS) + 1)] == [
        "DSC2204",
        "Lecture",
        1,
        "All",
        "Wed",
        "0900",
        "1100",
        80,
        "PUNGGOL",
        None,
        "LECT-01",
        None,
        None,
        "DR TAN",
        None,
        "1,2,3",
        "A0",
        "w ESE1109",
    ]


def test_csv_export_uses_system_template_columns_and_values(db_session):
    run = _add_scheduled_session(db_session)

    csv_text = ExportService().csv_buffer(db_session, run.id).getvalue()
    rows = list(csv.DictReader(StringIO(csv_text)))

    assert csv_text.splitlines()[0] == ",".join(SYSTEM_TEMPLATE_COLUMNS)
    assert rows[0]["Module"] == "DSC2204"
    assert rows[0]["Class Type"] == "Lecture"
    assert rows[0]["Group"] == "All"
    assert rows[0]["Day"] == "Wed"
    assert rows[0]["Start"] == "0900"
    assert rows[0]["End"] == "1100"
    assert rows[0]["Staff1"] == "DR TAN"
    assert rows[0]["Tri Week"] == "1,2,3"
    assert rows[0]["Recording Mode"] == "A0"
    assert rows[0]["Remark"] == "w ESE1109"
